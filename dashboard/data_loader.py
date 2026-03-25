from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook

BASE_DIR = Path.home() / "Desktop" / "全自动运行表格保存"
STATUS_MAP = {
    "增量昵称简介": "⚡",
    "TapTap-头像-用户资料图片": "✓",
    "TapTap-融媒体长文本": "⚠️",
    "steam昵称简介": "⚠️",
    "steam头像封面": "✓",
    "steam成就标题简介": "🔴",
    "战绩昵称": "🔴",
    "战绩头像": "✓",
    "海外小镇书籍": "✓",
    "国内小镇书籍": "✓",
    "国内小镇舆情": "⚠️",
    "海外小镇舆情": "⚠️",
    "国内小镇照片": "⚡",
    "海外小镇照片": "✓",
    "融媒体短文本": "⚠️",
}
BIZ_GROUP_MAP = {
    "steam成就标题简介": "Steam",
    "steam昵称简介": "Steam",
    "steam头像封面": "Steam",
    "TapTap-头像-用户资料图片": "TapTap",
    "TapTap-融媒体长文本": "TapTap",
    "增量昵称简介": "昵称资料",
    "战绩昵称": "战绩",
    "战绩头像": "战绩",
    "融媒体短文本": "融媒体",
    "国内小镇书籍": "国内小镇",
    "国内小镇照片": "国内小镇",
    "国内小镇舆情": "国内小镇",
    "海外小镇书籍": "海外小镇",
    "海外小镇照片": "海外小镇",
    "海外小镇舆情": "海外小镇",
}
HEADER_ALIASES = {
    "date": ["日期"],
    "total_count": ["总进审量", "1"],
    "push_rate": ["人工推审率", "人审推审率", "推审率(手动)", "初审推审率", "推审率"],
    "violation_rate": ["大盘违规率", "违规率(手动)", "总违规率", "违规率"],
    "review_count": ["大模型进审量", "人审总量", "小模型均通过量"],
    "reject_count": ["大模型人审驳回量", "驳回量", "总驳回量"],
    "error_value": ["数据量误差值", "小模型漏放黑样本数"],
}


@dataclass
class MetricRow:
    date: date
    biz_name: str
    total_count: Optional[float]
    push_rate: Optional[float]
    violation_rate: Optional[float]
    review_count: Optional[float]
    reject_count: Optional[float]
    error_value: Optional[float]
    base_status: str
    source_file: str


def parse_date(raw) -> Optional[date]:
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    if raw in (None, ""):
        return None
    s = str(raw).strip().replace("/", "-")
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def safe_float(v) -> Optional[float]:
    return float(v) if isinstance(v, (int, float)) else None


def first_header_index(headers: list[str | None], aliases: list[str]) -> Optional[int]:
    for alias in aliases:
        for idx, header in enumerate(headers):
            if header == alias:
                return idx
    return None


def extract_field(row: list | tuple, idx: Optional[int]) -> Optional[float]:
    if idx is None or idx >= len(row):
        return None
    return safe_float(row[idx])


def header_index_map(ws) -> dict[str, Optional[int]]:
    headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    return {key: first_header_index(headers, aliases) for key, aliases in HEADER_ALIASES.items()}


def load_metrics(base_dir: Path = BASE_DIR) -> pd.DataFrame:
    rows: list[MetricRow] = []

    for p in sorted(base_dir.glob("*.xlsx")):
        biz = p.stem.split("_", 1)[1]
        wb = load_workbook(p, data_only=False, read_only=True)
        ws = wb[wb.sheetnames[0]]
        idx_map = header_index_map(ws)

        for row in ws.iter_rows(min_row=2, values_only=True):
            d = parse_date(row[idx_map["date"]] if idx_map["date"] is not None else None)
            if not d:
                continue
            rows.append(
                MetricRow(
                    date=d,
                    biz_name=biz,
                    total_count=extract_field(row, idx_map["total_count"]),
                    push_rate=extract_field(row, idx_map["push_rate"]),
                    violation_rate=extract_field(row, idx_map["violation_rate"]),
                    review_count=extract_field(row, idx_map["review_count"]),
                    reject_count=extract_field(row, idx_map["reject_count"]),
                    error_value=extract_field(row, idx_map["error_value"]),
                    base_status=STATUS_MAP.get(biz, "✓"),
                    source_file=str(p.name),
                )
            )

    df = pd.DataFrame([r.__dict__ for r in rows])
    if df.empty:
        return df

    df["date"] = pd.to_datetime(df["date"])
    df["day"] = df["date"].dt.date
    df["biz_group"] = df["biz_name"].map(BIZ_GROUP_MAP).fillna("其他")
    df = df.sort_values(["biz_name", "date"]).reset_index(drop=True)

    df["push_rate_pct"] = df["push_rate"] * 100
    df["violation_rate_pct"] = df["violation_rate"] * 100
    df["total_7d_avg"] = df.groupby("biz_name")["total_count"].transform(lambda s: s.rolling(7, min_periods=1).mean())
    df["push_rate_7d_avg"] = df.groupby("biz_name")["push_rate"].transform(lambda s: s.rolling(7, min_periods=1).mean())
    df["violation_rate_7d_avg"] = df.groupby("biz_name")["violation_rate"].transform(lambda s: s.rolling(7, min_periods=1).mean())
    df["review_7d_avg"] = df.groupby("biz_name")["review_count"].transform(lambda s: s.rolling(7, min_periods=1).mean())

    df["total_vs_7d_pct"] = (df["total_count"] - df["total_7d_avg"]) / df["total_7d_avg"]
    df["push_vs_7d_pct"] = (df["push_rate"] - df["push_rate_7d_avg"]) / df["push_rate_7d_avg"]
    df["vio_vs_7d_pct"] = (df["violation_rate"] - df["violation_rate_7d_avg"]) / df["violation_rate_7d_avg"]
    return df


def latest_snapshot(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    global_latest = df["date"].max()
    snap = df.sort_values(["biz_name", "date"]).groupby("biz_name", as_index=False).tail(1).copy()
    snap["freshness_days"] = (global_latest - snap["date"]).dt.days
    snap["status"] = snap["base_status"]
    snap.loc[snap["freshness_days"] > 0, "status"] = "🔴"
    snap["is_abnormal"] = snap["status"].isin(["🔴", "⚠️", "⚡"])
    snap["freshness_label"] = snap["freshness_days"].map(lambda x: "当日" if x == 0 else f"滞后{x}天")
    return snap.sort_values(["status", "biz_group", "biz_name"]).reset_index(drop=True)
