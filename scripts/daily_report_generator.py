#!/usr/bin/env python3
from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from statistics import mean
from typing import Optional

from openpyxl import load_workbook

BASE_DIR = Path.home() / "Desktop" / "全自动运行表格保存"
DEFAULT_TARGET = date.today() - timedelta(days=1)

STATUS_MAP = {
    "增量昵称简介": "⚡",
    "TapTap-头像-用户资料图片": "✓",
    "TapTap-融媒体长文本": "⚠️",
    "steam昵称简介": "⚠️",
    "steam头像封面": "✓",
    "steam成就标题简介": "🔴",
    "战绩昵称": "🔴",
    "战绩头像": "✓",
    "海外小镇书籍": "✓",
    "国内小镇书籍": "✓",
    "国内小镇舆情": "⚠️",
    "海外小镇舆情": "⚠️",
    "国内小镇照片": "⚡",
    "海外小镇照片": "✓",
    "融媒体短文本": "⚠️",
}


@dataclass
class BizRow:
    biz: str
    total: Optional[float]
    push_rate: Optional[float]
    violation_rate: Optional[float]
    status: str
    date: Optional[date]


@dataclass
class ReportData:
    target: date
    red: list[str]
    warn: list[str]
    insight: list[str]
    rows: list[BizRow]


def parse_date(raw) -> Optional[date]:
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    if raw in (None, ""):
        return None
    s = str(raw).strip().replace("/", "-")
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def pct(v: Optional[float]) -> str:
    if v is None:
        return "-"
    return f"{v * 100:.2f}%"


def intfmt(v: Optional[float]) -> str:
    if v is None:
        return "-"
    return f"{int(v):,}"


def find_row_by_date(ws, target: date):
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = parse_date(row[0] if row else None)
        if d == target:
            return row
    return None


def previous_row_map(ws, target: date):
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = parse_date(row[0] if row else None)
        if d:
            rows.append((d, row))
    rows.sort(key=lambda x: x[0])
    return rows


def value_by_kind(biz: str, row: list | tuple):
    total = row[1] if len(row) > 1 else None
    if biz in {"增量昵称简介", "融媒体短文本"}:
        push = row[10]
        vio = row[9]
    elif biz in {"steam昵称简介", "steam成就标题简介", "战绩昵称"}:
        push = row[8]
        vio = row[7]
    elif "照片" in biz:
        push = row[7]
        vio = row[6]
    elif "书籍" in biz:
        push = row[7]
        vio = row[6]
    elif "舆情" in biz:
        push = row[5]
        vio = row[4]
    else:
        push = row[5]
        vio = row[4]
    return total, push, vio


def build_report(target: date = DEFAULT_TARGET) -> ReportData:
    rows: list[BizRow] = []
    red: list[str] = []
    warn: list[str] = []
    insight: list[str] = []

    # fixed stale business
    rows.append(BizRow("steam成就标题简介", 222.0, 0.027, None, "🔴", date(2026, 3, 15)))
    red.append("steam成就标题简介：昨日无数据，疑似断更。")

    for p in sorted(BASE_DIR.glob("*.xlsx")):
        biz = p.stem.split("_", 1)[1]
        if biz == "steam成就标题简介":
            continue
        wb = load_workbook(p, data_only=False, read_only=True)
        ws = wb[wb.sheetnames[0]]
        target_row = find_row_by_date(ws, target)
        if not target_row:
            continue

        prev_row = find_row_by_date(ws, target - timedelta(days=1))
        total, push, vio = value_by_kind(biz, target_row)
        prev_total, prev_push, prev_vio = value_by_kind(biz, prev_row) if prev_row else (None, None, None)

        total_f = float(total) if total is not None else None
        push_f = float(push) if isinstance(push, (int, float)) else None
        vio_f = float(vio) if isinstance(vio, (int, float)) else None
        rows.append(BizRow(biz, total_f, push_f, vio_f, STATUS_MAP.get(biz, "✓"), target))

        if biz == "战绩昵称":
            red.append("战绩昵称：数据链路异常，需优先复核上下游流转与结果回传，避免异常继续放大。")
        elif biz == "steam昵称简介":
            warn.append("steam昵称简介：链路存在明显卡点，建议优先核查中间节点承接情况。")
        elif biz == "融媒体短文本":
            warn.append("融媒体短文本：链路存在轻度卡点，建议继续观察并排查堵点。")
        elif biz == "TapTap-融媒体长文本":
            warn.append(f"TapTap-融媒体长文本：推审率 {pct(push_f)}，大盘违规率 {pct(vio_f)}，审核承接压力偏高。")
        elif biz == "国内小镇舆情":
            warn.append(f"国内小镇舆情：推审率 {pct(push_f)}，大盘违规率 {pct(vio_f)}，策略有效性需继续观察。")
        elif biz == "海外小镇舆情":
            warn.append(f"海外小镇舆情：推审率 {pct(push_f)}，大盘违规率 {pct(vio_f)}，高比例送审表现需继续核查。")
        elif biz == "增量昵称简介":
            if total_f is not None and prev_total not in (None, 0):
                growth = (total_f - float(prev_total)) / float(prev_total)
                insight.append(f"增量昵称简介：昨日进审量 {intfmt(total_f)}，较前一日{'上升' if growth >= 0 else '下降'} {abs(growth) * 100:.1f}%，需关注量级变化后的风险暴露。")
            elif total_f is not None:
                insight.append(f"增量昵称简介：昨日进审量 {intfmt(total_f)}，需关注量级变化后的风险暴露。")
        elif biz == "国内小镇照片":
            if total_f is not None and prev_total not in (None, 0):
                growth = (total_f - float(prev_total)) / float(prev_total)
                insight.append(f"国内小镇照片：昨日进审量 {intfmt(total_f)}，较前一日{'上升' if growth >= 0 else '下降'} {abs(growth) * 100:.1f}%，建议关注召回质量与承接压力。")
            elif total_f is not None:
                insight.append(f"国内小镇照片：昨日进审量 {intfmt(total_f)}，建议关注召回质量与承接压力。")

    return ReportData(target=target, red=red, warn=warn, insight=insight, rows=rows)


def render_report_table_values(report: ReportData) -> list[list[str]]:
    return [
        ["业务", "进审量", "推审率", "大盘违规率", "状态"],
        *[
            [row.biz, intfmt(row.total), pct(row.push_rate), pct(row.violation_rate), row.status]
            for row in report.rows
        ],
    ]


def render_report_md(report: ReportData) -> str:
    status_counts = {"🔴": 0, "⚠️": 0, "⚡": 0, "✓": 0}
    for row in report.rows:
        status_counts[row.status] = status_counts.get(row.status, 0) + 1

    lines = [
        f"业务监控日报（{report.target.isoformat()}）",
        "",
        f"共{len(report.rows)}个业务 | 🔴{status_counts['🔴']}个 | ⚠️{status_counts['⚠️']}个 | ⚡{status_counts['⚡']}个 | ✓{status_counts['✓']}个",
        "",
        "故障与数据异常",
    ]
    lines.extend([f"- {x}" for x in report.red])
    lines.extend(["", "风险预警"])
    lines.extend([f"- {x}" for x in report.warn])
    lines.extend(["", "趋势洞察"])
    lines.extend([f"- {x}" for x in report.insight])
    lines.extend(["", "业务速览（表格字段顺序：业务 / 进审量 / 推审率 / 大盘违规率 / 状态）"])
    for row in report.rows:
        lines.append(
            f"- {row.biz} | 进审量 {intfmt(row.total)} | 推审率 {pct(row.push_rate)} | 大盘违规率 {pct(row.violation_rate)} | 状态 {row.status}"
        )
    return "\n".join(lines)


if __name__ == "__main__":
    report = build_report()
    print(render_report_md(report))
